"""
app.py — Streamlit web app for converting supplier PDF quotes into two
Excel files: a Purchase Order Lines file and a Catalogue Lines file, both
formatted for system import. Supports Haymans, Cetnaj, Ideal Electrical,
Process Systems (valvesonline.com.au), APS Industrial, IPD Group, Phoenix
Contact, Mechtric, NHP, and Dore Electrics layouts — auto-detected from
PDF content. An existing Purchase Order Lines .xlsx can also be uploaded
directly to re-generate its Catalogue Lines file.
"""

import base64
import hashlib
import io
import re
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from quote_to_excel import parse_quote_pdf

# Storage is optional and must never take the app down. If the module or its
# dependency (supabase) is missing, logging is silently disabled and the
# converter keeps working.
try:
    import supabase_store
except Exception:
    supabase_store = None


# --- Constants matching the Purchase Order Lines template ----------------

PO_HEADERS = [
    "Job Code\n(*text 20)",
    "Part No on catalogue line\n(text 20)",
    "Activity Code\n(*text 10)",
    "Work Centre\n(*text 10)",
    "Description\n(text 100)",
    "Details\n(text 2000)",
    "Quantity\n(*number)",
    "Unit\n(text 10)",
    "Unit Cost\n(*number)",
]
PO_COLUMN_WIDTHS = {
    "A": 15.14, "B": 19.0, "C": 10.29, "D": 9.86,
    "E": 58.86, "F": 4.0,  "G": 10.71, "H": 7.71, "I": 10.71,
}

# --- Constants matching the Catalogue Lines template --------------------

CATALOGUE_HEADERS = [
    "Catalogue Part No\n(text 20)",
    "Supplier Part No\n(text 30)",
    "Description\n(*text 100)",
    "Activity Code\n(*text 10)",
    "Specification\n(text 2000)",
    "Favourite\n(integer)",
    "Promotional\n(integer)",
    "Cost Rate\n(number)",
    "Sell Rate\n(number)",
    "Unit\n(text 10)",
    "Negotiated % Disct\n(number)",
    "Negotiated Date\n(date)",
    "Last Invoice No\n(text 10)",
    "Last Invoice Cost\n(number)",
    "Negotiated Date\n(date)",
    "Bill Type\n(text 2)",
    "Group Code\n(text 10)",
    "SubCategory Code\n(text 10)",
    "Category Code\n(text 10)",
    "Inventory Code\n(text 20)",
    "Inactive\n(integer)",
    "Supplier\n(integer)",
    "Currency Code\n(*text 10)",
]
CATALOGUE_COLUMN_WIDTHS = {
    "A": 20.57, "B": 20.71, "C": 58.86, "D": 10.29, "E": 40.71,
    "F": 10.71, "G": 10.71, "H": 10.71, "I": 10.71, "J": 7.71,
    "K": 14.71, "L": 12.71, "M": 11.71, "N": 12.71, "O": 12.71,
    "P": 7.71,  "Q": 9.14,  "R": 13.86, "S": 11.14, "T": 11.57,
    "U": 10.71, "V": 10.71, "W": 11.0,
}
DEFAULT_CURRENCY = "AUD"

HEADER_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
DEFAULT_WORK_CENTRE = "WC004"

# Suppliers whose part numbers carry a 3-character prefix that should be
# stripped to produce the "Supplier Part No" (column B in the catalogue).
# Example: "SPFCTR-60C-5R" → "CTR-60C-5R" (Haymans/Cetnaj convention).
_PREFIX_STRIP_SUPPLIERS = {"haymans", "cetnaj", "ideal"}

# Pattern matching freight-charge codes for Haymans / Cetnaj, e.g.
# "SPF-FREIGHT", "BUN-FREIGHT", "APS-FREIGHT". These look like real part
# numbers but they're supplier-specific freight line items, not stocked
# parts, so we leave column B blank for them.
_FREIGHT_PART_RE = re.compile(r"^[A-Z0-9]{1,4}-FREIGHT$", re.IGNORECASE)

# Description fields in the target system have a 100-character cap. Truncate
# any longer description before writing it to either spreadsheet so imports
# don't get rejected. Applied to both PO and Catalogue output.
DESCRIPTION_MAX_LEN = 100


def _truncate_description(desc: str | None) -> str | None:
    """Trim a description to fit the target system's character limit.

    Returns None unchanged (so blank cells stay blank). Uses a hard cut at
    DESCRIPTION_MAX_LEN — we don't try to be clever about word boundaries
    since the field is for import, not for human reading.
    """
    if desc is None:
        return None
    if len(desc) <= DESCRIPTION_MAX_LEN:
        return desc
    return desc[:DESCRIPTION_MAX_LEN]


def _supplier_part_no(part: str, supplier: str | None) -> str:
    """Return the value to put in 'Supplier Part No' (catalogue column B).

    See the docstring on build_catalogue_xlsx for the full rule list.
    """
    if supplier not in _PREFIX_STRIP_SUPPLIERS:
        return ""
    # Haymans / Cetnaj freight codes like "SPF-FREIGHT" stay blank.
    if supplier in ("haymans", "cetnaj") and _FREIGHT_PART_RE.match(part):
        return ""
    if len(part) <= 3:
        return ""
    return part[3:]


# --- Page setup ----------------------------------------------------------

st.set_page_config(
    page_title="Quote PDF → Purchase Order",
    page_icon="📄",
    layout="centered",
)

st.title("📄 Quote PDF → Purchase Order")
st.write(
    "Upload a supplier PDF quote. "
    "The app extracts line items and fills the Purchase Order Lines template."
)


# --- Cached PDF parsing --------------------------------------------------

@st.cache_data(show_spinner=False)
def parse_pdf_bytes(pdf_bytes: bytes, _file_hash: str):
    """Parse a supplier quote PDF from raw bytes.

    Detection of supplier and routing to the right parser happens in
    quote_to_excel.parse_quote_pdf. We just pass the bytes through.
    """
    return parse_quote_pdf(pdf_bytes)


# --- Spreadsheet (.xlsx) inputs → items ----------------------------------
#
# Two spreadsheet shapes are supported in addition to supplier PDFs:
#   * "po_lines"            — an existing Purchase Order Lines export, re-
#                             converted into a Catalogue file.
#   * "price_availability"  — a supplier Price & Availability export.
# parse_xlsx() sniffs the header row and routes to the right reader. Both
# return the same item-dict shape the PDF parsers produce, so the preview,
# subtotal, PO and Catalogue builds downstream are unchanged.

# Purchase Order Lines layout — 0-based column positions matching PO_HEADERS.
_PO_COL_PART = 1      # B — Part No on catalogue line
_PO_COL_ACTIVITY = 2  # C — Activity Code
_PO_COL_DESC = 4      # E — Description
_PO_COL_QTY = 6       # G — Quantity
_PO_COL_UNIT = 7      # H — Unit
_PO_COL_COST = 8      # I — Unit Cost


def _xlsx_clean(v):
    """Trim a cell to a non-empty string, or None."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# Unit-of-measure codes that mean "each" and normalise to EA (the system's
# convention). Anything else is passed through uppercased so genuine units
# like M (metre) or KG survive.
_EACH_UOMS = {"EA", "EACH", "PCE", "PC", "PCS", "PIECE", "NO", "UN", "UNIT"}


def _normalise_uom(uom):
    if not uom:
        return None
    u = str(uom).strip().upper()
    return "EA" if u in _EACH_UOMS else (u or None)


def _to_int_qty(v):
    try:
        return int(float(v)) if v not in (None, "") else 1
    except (TypeError, ValueError):
        return 1


def _parse_po_lines_sheet(ws):
    """Read a Purchase Order Lines sheet into items.

    Unit Cost in the PO template is already a per-single-unit price, so
    unit_price is taken as-is with per == 1. Each row carries its own Activity
    Code (column C) so the catalogue keeps it per-row. Rows without a part
    number or a numeric cost (blanks, totals, footers) are skipped.
    """
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row) + [None] * (9 - len(row))
        part = _xlsx_clean(cells[_PO_COL_PART])
        cost = cells[_PO_COL_COST]
        if part is None or cost in (None, ""):
            continue
        try:
            unit_price = float(cost)
        except (TypeError, ValueError):
            continue
        items.append({
            "part": part,
            "description": _xlsx_clean(cells[_PO_COL_DESC]),
            "qty": _to_int_qty(cells[_PO_COL_QTY]),
            "uom": _normalise_uom(_xlsx_clean(cells[_PO_COL_UNIT])),
            "unit_price": unit_price,
            "per": 1,
            "supplier": None,
            "activity_code": _xlsx_clean(cells[_PO_COL_ACTIVITY]),
        })
    return items


# Price & Availability exports in this layout are Schneider's website
# download. Tagging the supplier keeps the stored metadata consistent; it
# does NOT trigger the Haymans/Cetnaj/Ideal prefix-strip, so catalogue
# column B stays blank as before.
_PA_SUPPLIER = "schneider"


def _parse_price_availability_sheet(ws, headers):
    """Read a supplier Price & Availability export into items.

    Columns are matched by header name (robust to re-ordering). 'Unit Net
    Price' is the ex-GST per-unit cost; 'Qty' is the line quantity. Rows
    without a part number or a numeric net price are skipped.
    """
    idx = {h.strip().lower(): i for i, h in enumerate(headers) if h}

    def col(*names):
        for n in names:
            if n.lower() in idx:
                return idx[n.lower()]
        return None

    c_part = col("Part No.", "Part No", "Part Number")
    c_desc = col("Description")
    c_qty = col("Qty", "Quantity")
    c_price = col("Unit Net Price", "Net Price")
    c_uom = col("Unit of Measure", "UOM", "Unit")

    def g(cells, i):
        return cells[i] if (i is not None and i < len(cells)) else None

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row)
        part = _xlsx_clean(g(cells, c_part))
        price = g(cells, c_price)
        if part is None or price in (None, ""):
            continue
        try:
            unit_price = float(price)
        except (TypeError, ValueError):
            continue
        items.append({
            "part": part,
            "description": _xlsx_clean(g(cells, c_desc)),
            "qty": _to_int_qty(g(cells, c_qty)),
            "uom": _normalise_uom(_xlsx_clean(g(cells, c_uom))),
            "unit_price": unit_price,
            "per": 1,
            "supplier": _PA_SUPPLIER,
        })
    return items


def parse_xlsx(xlsx_bytes: bytes):
    """Detect the spreadsheet shape and parse it. Returns (kind, items)."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [(str(c.value).strip() if c.value is not None else "")
               for c in ws[1]]
    hset = {h.lower() for h in headers}
    # Price & Availability exports carry a 'Unit Net Price' column; the PO
    # Lines template does not. That's the distinguishing signature.
    if "unit net price" in hset and "part no." in hset:
        return "price_availability", _parse_price_availability_sheet(ws, headers)
    return "po_lines", _parse_po_lines_sheet(ws)


@st.cache_data(show_spinner=False)
def parse_xlsx_bytes(xlsx_bytes: bytes, _file_hash: str):
    """Cached wrapper mirroring parse_pdf_bytes, for .xlsx uploads."""
    return parse_xlsx(xlsx_bytes)


def build_po_xlsx(items, job_code: str, activity_code: str) -> bytes:
    """Write items into the Purchase Order Lines template layout in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row — matches the template's blue fill + white bold font
    ws.append(PO_HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 45

    # Data rows
    for it in items:
        # Haymans quotes the price as "Unit Price per N units" (e.g. $479 per
        # 100m). Divide by 'per' so the value written into the template is the
        # true per-single-unit cost. For the vast majority of rows per == 1,
        # so this is a no-op.
        unit_cost = it["unit_price"] / it["per"]
        ws.append([
            job_code or None,                          # A — Job Code
            it["part"] or None,                        # B — Part No (None for freight)
            activity_code or None,                     # C — Activity Code
            DEFAULT_WORK_CENTRE,                       # D — Work Centre
            _truncate_description(it["description"]),  # E — Description (100 char cap)
            None,                                      # F — Details (left blank)
            it["qty"],                                 # G — Quantity
            it["uom"] or None,                         # H — Unit (None for freight)
            unit_cost,                                 # I — Unit Cost
        ])

    # Column widths from the template
    for col, width in PO_COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_catalogue_xlsx(items, activity_code: str) -> bytes:
    """Write items into the Catalogue Lines template layout in memory.

    Columns filled per item:
      A — Catalogue Part No  (raw part number, all suppliers)
      B — Supplier Part No   (rules vary by supplier — see below)
      C — Description
      D — Activity Code      (optional, from UI)
      H — Cost Rate          (= unit_price / per — true per-unit cost)
      J — Unit               (UOM)
      W — Currency Code      (always 'AUD')

    Column B rules:
      * Haymans / Cetnaj: first 3 chars of part number stripped (e.g.
        "SPFCTR-60C-5R" → "CTR-60C-5R"). Exception: freight-charge codes
        of the form "XXX-FREIGHT" leave column B blank — supplier-specific
        freight charges aren't real catalogue parts.
      * Ideal: first 3 chars stripped. Ideal freight rows have no part
        number at all (they're "Charges Freight ..." lines) so they're
        already excluded earlier by the "if not part" check.
      * Other suppliers (unknown layout, fallback parser): column B blank.

    All other columns are left blank for the user to fill in if needed.
    Items without a part number (e.g. generic Haymans 'Freight' rows) are
    skipped entirely — the catalogue is for physical stocked items.

    Rows are de-duplicated by Catalogue Part No: a part appearing on multiple
    PO lines yields a single catalogue row (first occurrence wins).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row — same styling as the PO file
    ws.append(CATALOGUE_HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 45

    # Data rows — dedupe by part number. The same part can appear on several
    # PO lines (e.g. the same MCB used across multiple cells); the catalogue
    # holds one row per stocked part, so the first occurrence wins and later
    # repeats are skipped. No-part rows (freight etc.) are skipped entirely.
    seen_parts = set()
    for it in items:
        part = it["part"]
        if not part:
            # Skip rows with no part number (generic 'Freight' lines etc.)
            continue
        if part in seen_parts:
            continue
        seen_parts.add(part)
        unit_cost = it["unit_price"] / it["per"]
        supplier_part = _supplier_part_no(part, it.get("supplier"))
        row = [None] * 23
        row[0] = part                                            # A — Catalogue Part No
        row[1] = supplier_part or None                           # B — Supplier Part No
        row[2] = _truncate_description(it["description"]) or None  # C — Description (100 char cap)
        row[3] = it.get("activity_code") or activity_code or None  # D — Activity Code
        row[7] = unit_cost                                       # H — Cost Rate
        row[9] = it["uom"] or None                               # J — Unit
        row[22] = DEFAULT_CURRENCY                               # W — Currency Code
        ws.append(row)

    # Column widths
    for col, width in CATALOGUE_COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- UI ------------------------------------------------------------------

col1, col2 = st.columns(2)
with col1:
    job_code = st.text_input(
        "Job Code (optional)",
        value="",
        max_chars=20,
        help="Leave blank if you'd rather fill it in Excel.",
    )
with col2:
    activity_code = st.text_input(
        "Activity Code (optional)",
        value="",
        max_chars=10,
        help="Leave blank if you'd rather fill it in Excel.",
    )

uploaded = st.file_uploader(
    "Choose a quote PDF",
    type=["pdf", "xlsx"],
    accept_multiple_files=False,
)

if uploaded is None:
    st.info("👆 Drag a PDF here or click to browse.")
    st.stop()


pdf_bytes = uploaded.getvalue()
file_hash = hashlib.sha1(pdf_bytes).hexdigest()
# An uploaded .xlsx is sniffed and routed (PO Lines vs Price & Availability);
# anything else is parsed as a supplier PDF.
is_xlsx = Path(uploaded.name).suffix.lower() in (".xlsx", ".xlsm")

with st.spinner("Reading PDF…"):
    try:
        if is_xlsx:
            kind, items = parse_xlsx_bytes(pdf_bytes, file_hash)
        else:
            kind, items = "pdf", parse_pdf_bytes(pdf_bytes, file_hash)
    except Exception as exc:
        st.error(f"Couldn't parse that PDF: {exc}")
        st.stop()

if not items:
    st.warning(
        "No line items found. "
        "If this is from a new supplier with a different layout, "
        "the parser may need new rules."
    )
    st.stop()


st.success(f"Found **{len(items)}** line item(s).")

# Subtotal — lets the user cross-check against the PDF's 'Total Excl GST'
subtotal = sum(it["qty"] * it["unit_price"] / it["per"] for it in items)
col_a, col_b = st.columns(2)
col_a.metric("Line items", len(items))
col_b.metric("Subtotal (excl GST)", f"${subtotal:,.2f}")
st.caption(
    "💡 Compare the subtotal above with the **Total Excl GST** at the bottom "
    "of the PDF to confirm every line was captured correctly."
)

# Show a preview of what'll go into the template
st.dataframe(
    [
        {
            "Job Code": job_code or "",
            "Part No": it["part"],
            "Activity Code": activity_code or "",
            "Work Centre": DEFAULT_WORK_CENTRE,
            "Description": _truncate_description(it["description"]),
            "Quantity": it["qty"],
            "Unit": it["uom"],
            "Unit Cost": it["unit_price"] / it["per"],
        }
        for it in items
    ],
    use_container_width=True,
    hide_index=True,
)

# Build both xlsx files in memory. A PO Lines upload is passed straight
# through as the PO download (lossless) and only the Catalogue is derived;
# every other input (PDF or Price & Availability) builds a fresh PO from the
# parsed items.
if kind == "po_lines":
    po_xlsx_bytes = pdf_bytes
else:
    po_xlsx_bytes = build_po_xlsx(items, job_code.strip(), activity_code.strip())
catalogue_xlsx_bytes = build_catalogue_xlsx(items, activity_code.strip())

# Filenames mirror the input PDF
pdf_stem = Path(uploaded.name).stem
po_out_name = pdf_stem + "-PO.xlsx"
catalogue_out_name = pdf_stem + "-Catalogue.xlsx"

# --- Persist this conversion to private Supabase storage -----------------
#
# Logged once per uploaded file. Streamlit re-runs this whole script on every
# interaction (e.g. a download click), so we guard on file_hash in
# session_state to avoid duplicate rows. Storage is fail-soft: if Supabase
# isn't configured or a write fails, the app carries on and the user still
# gets their files. Verify saved records in the Supabase dashboard.
_logged = st.session_state.setdefault("_logged_hashes", set())
if (supabase_store is not None
        and file_hash not in _logged
        and supabase_store.is_configured()):
    saved = supabase_store.store_conversion(
        file_hash=file_hash,
        source_filename=uploaded.name,
        input_type="po_xlsx" if kind == "po_lines" else kind,
        supplier=(items[0].get("supplier") if items else None),
        job_code=job_code.strip(),
        activity_code=activity_code.strip(),
        line_count=len(items),
        subtotal_ex_gst=subtotal,
        catalogue_row_count=len({it["part"] for it in items if it.get("part")}),
        po_bytes=po_xlsx_bytes,
        po_name=po_out_name,
        catalogue_bytes=catalogue_xlsx_bytes,
        catalogue_name=catalogue_out_name,
    )
    if saved:
        _logged.add(file_hash)

# --- Single "Download both" button -------------------------------------
#
# Browsers limit a button click to one native download. We work around
# that by injecting a tiny HTML+JS component that contains both files
# as base64 data URLs and auto-clicks two hidden anchor tags — one for
# the PO file, one for the Catalogue file. Works in Chrome and Edge
# (and modern Firefox after a one-time "allow multiple downloads"
# prompt). If a browser blocks the second download silently, the user
# can fall back to the per-file links shown below the button.

if st.button(
    "⬇️  Download both files",
    type="primary",
    use_container_width=True,
):
    po_b64 = base64.b64encode(po_xlsx_bytes).decode()
    cat_b64 = base64.b64encode(catalogue_xlsx_bytes).decode()
    xlsx_mime = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # The hidden anchors get auto-clicked on render. A small delay between
    # the two clicks improves reliability in browsers that throttle rapid
    # programmatic downloads.
    components.html(
        f"""
        <html><body>
        <a id="dl1" href="data:{xlsx_mime};base64,{po_b64}"
           download="{po_out_name}" style="display:none">PO</a>
        <a id="dl2" href="data:{xlsx_mime};base64,{cat_b64}"
           download="{catalogue_out_name}" style="display:none">Catalogue</a>
        <script>
          document.getElementById('dl1').click();
          setTimeout(function() {{
            document.getElementById('dl2').click();
          }}, 400);
        </script>
        </body></html>
        """,
        height=0,
    )

# Fallback per-file links — small text below the button so the user
# always has a way to grab each file individually if their browser
# blocked the combined download or they just want one of the two.
fb_col1, fb_col2 = st.columns(2)
with fb_col1:
    st.download_button(
        label="PO only",
        data=po_xlsx_bytes,
        file_name=po_out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with fb_col2:
    st.download_button(
        label="Catalogue only",
        data=catalogue_xlsx_bytes,
        file_name=catalogue_out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

st.caption(
    f"Work Centre is set to **{DEFAULT_WORK_CENTRE}** for every row in the PO file. "
    "Job Code, Activity Code, and Details can be edited in Excel after download."
)
